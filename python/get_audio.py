import os
from TTS.api import TTS
import torch
from openpyxl import load_workbook

# Check if GPU is available
use_cuda = torch.cuda.is_available()
device = torch.device("cuda" if use_cuda else "cpu")
print(f"Using device: {device}")

# Initialize a global counter
download_counter = 0
errors = []

def retrieve_pronunciation(word, lang='fr'):
    global download_counter
    global errors
    directory = "C:/temp/"
    if not os.path.exists(directory):
        os.makedirs(directory)
    filename = f"{word}_{lang}.wav"
    filepath = os.path.join(directory, filename)

    download_counter += 1

    if not os.path.exists(filepath):
        try:
            # Use the French TTS model
            tts = TTS(model_name="tts_models/fr/css10/vits", progress_bar=True, gpu=use_cuda)
            tts.tts_to_file(text=word, file_path=filepath)
            print(f"Downloaded #{download_counter}: {filename}")
        except Exception as e:
            errors.append(word)
            print(f"#{download_counter} Error retrieving pronunciation for {word}: {e}")
        return None
    else:
        print(f"{download_counter}. File for {word} already exists. No download needed.")

    return filepath

def find_tables(worksheet):
    tables = []
    current_theme = None

    for row_cells in worksheet.iter_rows(min_row=1, max_col=2):
        values = [cell.value for cell in row_cells]
        if values[0] == "French" and values[1] == "English":
            start_row = row_cells[0].row
            end_row = start_row
            for following_row in worksheet.iter_rows(min_row=start_row + 1, max_col=2):
                following_values = [cell.value for cell in following_row]
                if not following_values[0] or following_values[0] == "French":
                    break
                end_row += 1
            tables.append((start_row, end_row, current_theme))
        else:
            if row_cells[1].value is None and isinstance(row_cells[0].value, str) and row_cells[0].value.strip() != "":
                current_theme = row_cells[0].value.strip()

    return tables

def process_tables(file_path):
    workbook = load_workbook(filename=file_path)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        table_ranges = find_tables(worksheet)

        for start_row, end_row, theme in table_ranges:
            data = worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=2, values_only=True)
            for french, english in data:
                if french and english and french != "French":
                    retrieve_pronunciation(french, 'fr')

if __name__ == "__main__":
    file_path = "C:/Developer/french-for-python.xlsx"
    process_tables(file_path)
    print(errors)
