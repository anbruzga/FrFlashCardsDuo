import base64
import csv
import os

from openpyxl import load_workbook


def load_pronunciation(filepath):
    """ Load an existing pronunciation file and return its base64 encoding. """
    if os.path.exists(filepath):
        with open(filepath, 'rb') as f:
            pronunciation_data = f.read()
        return base64.b64encode(pronunciation_data).decode('utf-8')
    return ''

def recreate_csv(file_path, csv_file_path='flashcards.csv'):
    cards = []
    workbook = load_workbook(filename=file_path)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        table_ranges = find_tables(worksheet)

        for start_row, end_row, theme in table_ranges:
            data = worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=2, values_only=True)
            for french, english in data:
                if french and english and french != "French":
                    filename = f"C:/temp/{french}_fr.wav"
                    french_pronunciation = load_pronunciation(filename)
                    card = {
                        'ID': len(cards) + 1,
                        'English': english,
                        'French': french,
                        'Theme': theme,
                        'French_Pronunciation': french_pronunciation
                    }
                    cards.append(card)

    save_to_csv(cards, csv_file_path)

def save_to_csv(flashcards, csv_file_path='flashcards.csv'):
    with open(csv_file_path, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(['ID', 'English', 'French', 'Theme', 'French_Pronunciation'])
        for card in flashcards:
            writer.writerow([card['ID'], card['English'], card['French'], card['Theme'], card['French_Pronunciation']])

def find_tables(worksheet):
    tables = []
    current_theme = None

    for row_cells in worksheet.iter_rows(min_row=1, max_col=2):
        values = [cell.value for cell in row_cells]
        if values[0] == "French" and values[1] == "English":
            start_row = row_cells[0].row
            end_row = start_row
            for following_row in worksheet.iter_rows(min_row=start_row + 1, max_col=2):
                following_values = [cell.value for cell in following_row]
                if not following_values[0] or following_values[0] == "French":
                    break
                end_row += 1
            tables.append((start_row, end_row, current_theme))
        else:
            if row_cells[1].value is None and isinstance(row_cells[0].value, str) and row_cells[0].value.strip() != "":
                current_theme = row_cells[0].value.strip()

    return tables

if __name__ == "__main__":
    file_path = "C:/Developer/french-for-python.xlsx"
    recreate_csv(file_path)
